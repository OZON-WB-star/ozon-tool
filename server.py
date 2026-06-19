from __future__ import annotations

import cgi
import json
import mimetypes
import socket
import sys
import threading
import time
import urllib.request
import uuid
import warnings
from dataclasses import asdict
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict, Iterable, List
from urllib.parse import urlparse

warnings.filterwarnings("ignore", category=DeprecationWarning, module=r"cgi")

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from openpyxl import load_workbook

from finder1688.excel_io import ExcelWorkbookAdapter
from finder1688.pipeline import FinderPipeline
from finder1688.playwright_finder import Playwright1688Finder

BASE_DIR = Path(__file__).resolve().parent
INDEX_FILE = BASE_DIR / "index.html"

DEFAULT_PROJECT_ROOT = Path(r"F:\OZON-PY\OZON_1688独立软件_final")
PROJECT_ROOT = DEFAULT_PROJECT_ROOT if DEFAULT_PROJECT_ROOT.exists() else ROOT_DIR
LOCAL_RUNTIME_ROOT = BASE_DIR / "data"


def _pick_existing_path(paths: Iterable[Path]) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    return None


def _ensure_writable_dir(preferred: Path, fallback: Path) -> Path:
    for path in (preferred, fallback):
        try:
            path.mkdir(parents=True, exist_ok=True)
            probe = path / ".write_test"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return path
        except Exception:
            continue
    raise RuntimeError(f"无法创建运行目录：{preferred} / {fallback}")


PRIMARY_EXPORT_DIR = (
    _pick_existing_path([PROJECT_ROOT / "EXPORT", PROJECT_ROOT / "ozon_core" / "EXPORT"])
    or (PROJECT_ROOT / "EXPORT")
)
DATA_DIR = _ensure_writable_dir(PROJECT_ROOT / "WEB_RUNTIME", LOCAL_RUNTIME_ROOT)
UPLOAD_DIR = _ensure_writable_dir(DATA_DIR / "uploads", LOCAL_RUNTIME_ROOT / "uploads")
DOWNLOAD_DIR = _ensure_writable_dir(PROJECT_ROOT / "downloads", LOCAL_RUNTIME_ROOT / "downloads")
DEBUG_DIR = _ensure_writable_dir(DOWNLOAD_DIR / "debug", LOCAL_RUNTIME_ROOT / "downloads" / "debug")
RUNTIME_LOG = DATA_DIR / "runtime.log"

EXPORT_DIRS = [
    PRIMARY_EXPORT_DIR,
    PROJECT_ROOT / "ozon_core" / "EXPORT",
    ROOT_DIR / "EXPORT",
    ROOT_DIR / "ozon_core" / "EXPORT",
    UPLOAD_DIR,
]

HOST = "127.0.0.1"
PORT = 8000

try:
    PRIMARY_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass


def log_runtime(message: str) -> None:
    stamp = time.strftime("%Y-%m-%d %H:%M:%S")
    with RUNTIME_LOG.open("a", encoding="utf-8") as handle:
        handle.write(f"[{stamp}] {message}\n")


STATE_LOCK = threading.Lock()
STATE: Dict[str, Any] = {
    "project_root": str(PROJECT_ROOT),
    "current_source": None,
    "current_source_path": None,
    "current_result": None,
    "current_result_path": None,
    "current_erp": None,
    "current_erp_path": None,
    "job": None,
    "message": "等待定位最新导出表。",
    "cdp_url": "http://127.0.0.1:9222",
}

JOB_CONTROL: Dict[str, Any] = {
    "thread": None,
    "stop_event": None,
    "kind": None,
}


def _set_state(**kwargs: Any) -> None:
    with STATE_LOCK:
        STATE.update(kwargs)


def _get_state() -> Dict[str, Any]:
    with STATE_LOCK:
        return dict(STATE)


def _safe_filename(name: str) -> str:
    return Path(name).name.replace("/", "_").replace("\\", "_")


def _human_size(size: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}"
        value /= 1024
    return f"{size} B"


def _recent_exports(limit: int = 8) -> List[Path]:
    candidates: list[Path] = []
    seen: set[Path] = set()
    for export_dir in EXPORT_DIRS:
        if not export_dir.exists():
            continue
        for item in export_dir.glob("*.xlsx"):
            resolved = item.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            candidates.append(item)
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return candidates[:limit]


def _latest_source_export() -> Path | None:
    candidates: list[Path] = []
    seen: set[Path] = set()
    for export_dir in EXPORT_DIRS:
        if not export_dir.exists():
            continue
        for item in export_dir.glob("*.xlsx"):
            if item.name.endswith("_1688_ready.xlsx") or item.name.endswith("_erp_ready.xlsx"):
                continue
            resolved = item.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            candidates.append(item)
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _default_1688_output(source_path: Path) -> Path:
    return PRIMARY_EXPORT_DIR / f"{source_path.stem}_1688_ready{source_path.suffix}"


def _default_erp_output(source_path: Path) -> Path:
    name = source_path.name
    if name.endswith("_1688_ready.xlsx"):
        return PRIMARY_EXPORT_DIR / name.replace("_1688_ready.xlsx", "_erp_ready.xlsx")
    return PRIMARY_EXPORT_DIR / f"{source_path.stem}_erp_ready{source_path.suffix}"


def _inspect_workbook(path: Path) -> Dict[str, Any]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return {
            "sheet_count": 0,
            "target_rows": 0,
            "file_size_text": _human_size(path.stat().st_size) if path.exists() else "0 B",
            "error": str(exc),
            "preview_rows": [],
        }

    try:
        adapter = ExcelWorkbookAdapter(path)
        preview_rows = list(adapter.iter_target_rows())[:8]
        target_rows = sum(1 for _ in adapter.iter_target_rows())
    except Exception:
        preview_rows = []
        target_rows = 0

    return {
        "sheet_count": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames,
        "target_rows": target_rows,
        "file_size_text": _human_size(path.stat().st_size),
        "preview_rows": [asdict(row) for row in preview_rows],
    }


def _active_preview_path() -> Path | None:
    state = _get_state()
    for key in ("current_result_path", "current_source_path"):
        value = state.get(key)
        if value:
            path = Path(value)
            if path.exists():
                return path
    return None


def _is_port_open(host: str, port: int) -> bool:
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(0.7)
    try:
        sock.connect((host, port))
        return True
    except Exception:
        return False
    finally:
        try:
            sock.close()
        except Exception:
            pass


def _check_browser(cdp_url: str) -> Dict[str, Any]:
    cdp_url = (cdp_url or "http://127.0.0.1:9222").strip()
    try:
        parsed = urlparse(cdp_url)
        host = parsed.hostname or "127.0.0.1"
        port = parsed.port or 9222
    except Exception:
        return {"ok": False, "error": f"浏览器调试地址格式不正确：{cdp_url}"}

    if not _is_port_open(host, port):
        return {"ok": False, "error": f"浏览器调试端口未连接：{cdp_url}"}

    try:
        with urllib.request.urlopen(f"{cdp_url}/json/version", timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8", errors="ignore"))
    except Exception as exc:
        return {"ok": False, "error": f"浏览器接口检测失败：{exc}"}

    browser_name = payload.get("Browser", "")
    ws_url = payload.get("webSocketDebuggerUrl", "")
    return {
        "ok": True,
        "browser": browser_name,
        "websocket": ws_url,
        "cdp_url": cdp_url,
    }


def _job_running() -> bool:
    state = _get_state()
    job = state.get("job") or {}
    return bool(job.get("running"))


def _start_background_job(kind: str, runner) -> Dict[str, Any]:
    if _job_running():
        return {"ok": False, "error": "已有任务正在运行，请先等待或停止当前任务。"}

    stop_event = threading.Event()
    job_id = uuid.uuid4().hex[:10]
    started_at = time.strftime("%Y-%m-%d %H:%M:%S")

    with STATE_LOCK:
        STATE["job"] = {
            "id": job_id,
            "kind": kind,
            "running": True,
            "status_text": "运行中",
            "started_at": started_at,
            "finished_at": "",
            "error": "",
            "processed": 0,
        }
        JOB_CONTROL["stop_event"] = stop_event
        JOB_CONTROL["kind"] = kind

    def worker() -> None:
        try:
            runner(stop_event)
        except Exception as exc:
            with STATE_LOCK:
                if STATE.get("job"):
                    STATE["job"] = {
                        **STATE["job"],
                        "running": False,
                        "status_text": "失败",
                        "error": str(exc),
                        "finished_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                    }
                STATE["message"] = f"{kind} 失败：{exc}"
            log_runtime(f"{kind} failed: {exc}")
        finally:
            with STATE_LOCK:
                JOB_CONTROL["thread"] = None
                JOB_CONTROL["stop_event"] = None
                JOB_CONTROL["kind"] = None

    thread = threading.Thread(target=worker, daemon=True)
    JOB_CONTROL["thread"] = thread
    thread.start()
    log_runtime(f"{kind} started")
    return {"ok": True, "job_id": job_id}


def _stop_current_job() -> Dict[str, Any]:
    stop_event = JOB_CONTROL.get("stop_event")
    if not stop_event:
        return {"ok": False, "error": "当前没有正在运行的任务。"}
    stop_event.set()
    _set_state(message="已发送停止指令，等待当前任务安全结束。")
    log_runtime("stop requested")
    return {"ok": True}


def _refresh_latest_export() -> Dict[str, Any]:
    latest = _latest_source_export()
    if not latest:
        _set_state(
            current_source=None,
            current_source_path=None,
            current_result=None,
            current_result_path=None,
            message="未找到新的 OZON 导出表。",
        )
        return {"ok": False, "error": "未找到新的 OZON 导出表。"}

    result_path = _default_1688_output(latest)
    erp_path = _default_erp_output(result_path)
    _set_state(
        current_source=latest.name,
        current_source_path=str(latest),
        current_result=result_path.name if result_path.exists() else None,
        current_result_path=str(result_path) if result_path.exists() else None,
        current_erp=erp_path.name if erp_path.exists() else None,
        current_erp_path=str(erp_path) if erp_path.exists() else None,
        message=f"已定位最新导出：{latest.name}",
    )
    log_runtime(f"latest export located: {latest}")
    return {"ok": True, "path": str(latest), "name": latest.name}


def _start_generate_erp() -> Dict[str, Any]:
    state = _get_state()
    source_value = state.get("current_result_path") or state.get("current_source_path")
    if not source_value:
        return {"ok": False, "error": "请先上传、刷新或运行 1688 结果表。"}

    source_path = Path(source_value)
    if not source_path.exists():
        return {"ok": False, "error": f"文件不存在：{source_path}"}

    output_path = _default_erp_output(source_path)

    def runner(stop_event: threading.Event) -> None:
        adapter = ExcelWorkbookAdapter(source_path)
        adapter.save(output_path)
        with STATE_LOCK:
            STATE["current_erp"] = output_path.name
            STATE["current_erp_path"] = str(output_path)
            STATE["job"] = {
                **STATE["job"],
                "running": False,
                "status_text": "完成",
                "processed": 1,
                "finished_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            }
            STATE["message"] = f"ERP 上品表已生成：{output_path.name}"
        log_runtime(f"erp generated: {output_path}")

    return _start_background_job("ERP生成", runner)


def _start_1688_processing(input_path: Path, output_path: Path, cdp_url: str) -> Dict[str, Any]:
    if not input_path.exists():
        return {"ok": False, "error": f"输入文件不存在：{input_path}"}

    browser = _check_browser(cdp_url)
    if not browser.get("ok"):
        return browser

    _set_state(cdp_url=cdp_url.strip())

    def runner(stop_event: threading.Event) -> None:
        finder = Playwright1688Finder(
            download_dir=DOWNLOAD_DIR,
            cdp_url=cdp_url.strip(),
            result_limit=3,
            stop_event=stop_event,
        )
        pipeline = FinderPipeline(input_path, finder, stop_event=stop_event)
        processed = pipeline.run(output_path)
        stopped = stop_event.is_set()
        with STATE_LOCK:
            STATE["current_source"] = input_path.name
            STATE["current_source_path"] = str(input_path)
            STATE["current_result"] = output_path.name
            STATE["current_result_path"] = str(output_path)
            STATE["job"] = {
                **STATE["job"],
                "running": False,
                "processed": processed,
                "status_text": "已停止" if stopped else "完成",
                "finished_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            }
            STATE["message"] = (
                f"1688 处理已停止，已处理 {processed} 行。"
                if stopped
                else f"1688 处理完成：{output_path.name}"
            )
        log_runtime(f"1688 finished: processed={processed}, stopped={stopped}, output={output_path}")

    return _start_background_job("1688处理", runner)


def _state_payload() -> Dict[str, Any]:
    state = _get_state()
    preview_path = _active_preview_path()
    summary = _inspect_workbook(preview_path) if preview_path else {}
    preview = summary.get("preview_rows", [])
    recent_files = []
    for path in _recent_exports():
        recent_files.append(
            {
                "name": path.name,
                "path": str(path),
                "mtime": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(path.stat().st_mtime)),
                "size": _human_size(path.stat().st_size),
            }
        )

    result_path = state.get("current_result_path")
    erp_path = state.get("current_erp_path")
    return {
        "project_root": state.get("project_root"),
        "export_root": str(PRIMARY_EXPORT_DIR),
        "current_source": state.get("current_source"),
        "current_source_path": state.get("current_source_path") or "",
        "current_result": state.get("current_result"),
        "current_result_path": result_path or "",
        "current_erp": state.get("current_erp"),
        "current_erp_path": erp_path or "",
        "message": state.get("message"),
        "job": state.get("job"),
        "summary": summary,
        "preview": preview,
        "recent_files": recent_files,
        "cdp_url": state.get("cdp_url"),
        "download_result_url": f"/download/{Path(result_path).name}" if result_path else "",
        "download_erp_url": f"/download/{Path(erp_path).name}" if erp_path else "",
    }


class WebHandler(BaseHTTPRequestHandler):
    server_version = "Ozon1688Web/2.0"

    def log_message(self, fmt: str, *args: Any) -> None:
        return

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self._send_html(_load_index_html())
            return
        if parsed.path == "/api/state":
            self._send_json(_state_payload())
            return
        if parsed.path.startswith("/download/"):
            self._send_download(parsed.path.removeprefix("/download/"))
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not Found")

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/upload":
            self._handle_upload()
            return
        if parsed.path == "/api/clear":
            self._handle_clear()
            return
        if parsed.path == "/api/refresh-latest":
            payload = _refresh_latest_export()
            self._send_json(payload, status=HTTPStatus.OK if payload.get("ok") else HTTPStatus.BAD_REQUEST)
            return
        if parsed.path == "/api/check-browser":
            body = self._read_json_body()
            cdp_url = body.get("cdp_url") or _get_state().get("cdp_url")
            payload = _check_browser(cdp_url)
            if payload.get("ok"):
                _set_state(cdp_url=cdp_url, message=f"浏览器检测通过：{payload.get('browser', '')}")
            self._send_json(payload, status=HTTPStatus.OK if payload.get("ok") else HTTPStatus.BAD_REQUEST)
            return
        if parsed.path == "/api/run-1688":
            body = self._read_json_body()
            state = _get_state()
            input_value = body.get("input_path") or state.get("current_source_path")
            cdp_url = body.get("cdp_url") or state.get("cdp_url")
            if not input_value:
                self._send_json({"ok": False, "error": "请先上传文件或刷新最新导出。"}, status=HTTPStatus.BAD_REQUEST)
                return
            input_path = Path(input_value)
            output_path = _default_1688_output(input_path)
            payload = _start_1688_processing(input_path, output_path, cdp_url)
            self._send_json(payload, status=HTTPStatus.OK if payload.get("ok") else HTTPStatus.BAD_REQUEST)
            return
        if parsed.path == "/api/stop-job":
            payload = _stop_current_job()
            self._send_json(payload, status=HTTPStatus.OK if payload.get("ok") else HTTPStatus.BAD_REQUEST)
            return
        if parsed.path == "/api/generate-erp":
            payload = _start_generate_erp()
            self._send_json(payload, status=HTTPStatus.OK if payload.get("ok") else HTTPStatus.BAD_REQUEST)
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not Found")

    def _read_json_body(self) -> Dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        if not raw:
            return {}
        try:
            return json.loads(raw.decode("utf-8"))
        except Exception:
            return {}

    def _handle_upload(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            self._send_json({"error": "请使用文件上传表单。"}, status=HTTPStatus.BAD_REQUEST)
            return

        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": content_type,
                "CONTENT_LENGTH": self.headers.get("Content-Length", ""),
            },
        )
        file_item = form["file"] if "file" in form else None
        if not file_item or not getattr(file_item, "filename", ""):
            self._send_json({"error": "没有收到上传文件。"}, status=HTTPStatus.BAD_REQUEST)
            return

        raw_name = _safe_filename(file_item.filename)
        if not raw_name.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            self._send_json({"error": "只支持 Excel 文件。"}, status=HTTPStatus.BAD_REQUEST)
            return

        upload_path = UPLOAD_DIR / f"{int(time.time())}_{raw_name}"
        with upload_path.open("wb") as handle:
            handle.write(file_item.file.read())

        _set_state(
            current_source=raw_name,
            current_source_path=str(upload_path),
            current_result=None,
            current_result_path=None,
            current_erp=None,
            current_erp_path=None,
            message=f"已上传文件：{raw_name}",
        )
        log_runtime(f"uploaded: {upload_path}")
        self._send_json({"ok": True, "filename": raw_name, "saved_path": str(upload_path)})

    def _handle_clear(self) -> None:
        if _job_running():
            self._send_json({"ok": False, "error": "请先停止当前任务。"}, status=HTTPStatus.BAD_REQUEST)
            return
        _set_state(
            current_source=None,
            current_source_path=None,
            current_result=None,
            current_result_path=None,
            current_erp=None,
            current_erp_path=None,
            job=None,
            message="当前工作状态已清空。",
        )
        log_runtime("state cleared")
        self._send_json({"ok": True})

    def _send_json(self, data: Dict[str, Any], status: int = HTTPStatus.OK) -> None:
        encoded = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def _send_html(self, text: str, status: int = HTTPStatus.OK) -> None:
        encoded = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def _send_download(self, filename: str) -> None:
        filename = _safe_filename(filename)
        candidate_paths = [
            PRIMARY_EXPORT_DIR / filename,
            UPLOAD_DIR / filename,
            DOWNLOAD_DIR / filename,
        ]
        for path in candidate_paths:
            if path.exists() and path.is_file():
                mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
                with path.open("rb") as handle:
                    blob = handle.read()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", mime)
                self.send_header("Content-Length", str(len(blob)))
                self.send_header("Content-Disposition", f'attachment; filename="{path.name}"')
                self.end_headers()
                self.wfile.write(blob)
                return
        self.send_error(HTTPStatus.NOT_FOUND, "File not found")


def _load_index_html() -> str:
    return INDEX_FILE.read_text(encoding="utf-8")


def main() -> None:
    _refresh_latest_export()
    server = ThreadingHTTPServer((HOST, PORT), WebHandler)
    print(f"Web app running at http://{HOST}:{PORT}", flush=True)
    log_runtime(f"server started on http://{HOST}:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
        log_runtime("server stopped")


if __name__ == "__main__":
    main()
